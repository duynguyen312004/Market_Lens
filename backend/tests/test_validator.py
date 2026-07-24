from io import BytesIO
from zipfile import ZIP_DEFLATED, ZipFile

import pandas as pd
import pytest
from openpyxl import load_workbook

from backend.app.core.errors import AppError
from backend.app.services import file_reader
from backend.app.services.file_reader import read_sales_file
from backend.app.services.validator import (
    REQUIRED_COLUMNS,
    validate_sales_data,
)


def valid_frame() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "order_id": "DH001",
                "order_date": "2026-07-01",
                "customer_id": "C001",
                "customer_name": "Nguyen Van A",
                "product_id": "P001",
                "product_name": "Ao thun",
                "category": "Thoi trang",
                "quantity": 2,
                "unit_price": 150_000,
                "discount": 20_000,
                "order_status": " Completed ",
            }
        ],
        columns=REQUIRED_COLUMNS,
    ).astype(object)


def assert_app_error(
    error: pytest.ExceptionInfo[AppError],
    code: str,
) -> AppError:
    assert error.value.code == code
    return error.value


def test_valid_data_is_normalized_and_revenue_is_calculated() -> None:
    result = validate_sales_data(valid_frame(), max_rows=50_000)

    assert result.loc[0, "order_status"] == "completed"
    assert result.loc[0, "quantity"] == 2
    assert result.loc[0, "line_revenue"] == 280_000
    assert result.loc[0, "order_date"].date().isoformat() == "2026-07-01"


def test_missing_column_returns_invalid_file_columns() -> None:
    frame = valid_frame().drop(columns=["discount"])

    with pytest.raises(AppError) as error:
        validate_sales_data(frame, max_rows=50_000)

    app_error = assert_app_error(error, "INVALID_FILE_COLUMNS")
    assert app_error.details["missing"] == ["discount"]


def test_normalized_duplicate_columns_are_rejected() -> None:
    frame = valid_frame()
    frame[" Discount "] = 0

    with pytest.raises(AppError) as error:
        validate_sales_data(frame, max_rows=50_000)

    app_error = assert_app_error(error, "INVALID_FILE_COLUMNS")
    assert app_error.details["duplicates"] == ["discount"]


@pytest.mark.parametrize(
    ("column", "value", "reason"),
    [
        ("order_date", "not-a-date", "invalid_date"),
        ("quantity", 0, "must_be_positive_integer"),
        ("quantity", -1, "must_be_positive_integer"),
        ("quantity", 1.5, "must_be_positive_integer"),
        ("unit_price", -1, "must_be_non_negative_number"),
        ("unit_price", float("inf"), "must_be_non_negative_number"),
        ("discount", -1, "must_be_non_negative_number"),
        ("order_status", "pending", "invalid_status"),
    ],
)
def test_invalid_row_values_return_structured_errors(
    column: str,
    value: object,
    reason: str,
) -> None:
    frame = valid_frame()
    frame.loc[0, column] = value

    with pytest.raises(AppError) as error:
        validate_sales_data(frame, max_rows=50_000)

    app_error = assert_app_error(error, "INVALID_ROW_DATA")
    assert {
        "row": 2,
        "column": column,
        "reason": reason,
    } in app_error.details["errors"]


def test_negative_line_revenue_is_rejected() -> None:
    frame = valid_frame()
    frame.loc[0, "discount"] = 400_000

    with pytest.raises(AppError) as error:
        validate_sales_data(frame, max_rows=50_000)

    app_error = assert_app_error(error, "INVALID_ROW_DATA")
    assert app_error.details["errors"][0]["reason"] == (
        "line_revenue_must_be_non_negative"
    )


def test_completely_blank_rows_are_ignored() -> None:
    blank = pd.DataFrame(
        [{column: None for column in REQUIRED_COLUMNS}],
        columns=REQUIRED_COLUMNS,
    )
    frame = pd.concat([valid_frame(), blank], ignore_index=True)

    result = validate_sales_data(frame, max_rows=50_000)

    assert len(result) == 1
    assert result.loc[0, "order_id"] == "DH001"


def test_row_error_response_is_capped_without_losing_total_count() -> None:
    frame = pd.concat([valid_frame()] * 25, ignore_index=True)
    frame["order_id"] = [f"DH{index:03d}" for index in range(25)]
    frame["quantity"] = 0
    frame["discount"] = 0

    with pytest.raises(AppError) as error:
        validate_sales_data(frame, max_rows=50_000)

    app_error = assert_app_error(error, "INVALID_ROW_DATA")
    assert len(app_error.details["errors"]) == 20
    assert app_error.details["total_error_count"] == 25


def test_conflicting_order_values_are_rejected_without_raw_row() -> None:
    frame = pd.concat([valid_frame(), valid_frame()], ignore_index=True)
    frame.loc[1, "customer_id"] = "C002"
    frame.loc[1, "customer_name"] = "Tran Thi B"

    with pytest.raises(AppError) as error:
        validate_sales_data(frame, max_rows=50_000)

    app_error = assert_app_error(error, "INVALID_ROW_DATA")
    conflict = next(
        item
        for item in app_error.details["errors"]
        if item.get("identifier") == "DH001"
    )
    assert conflict == {
        "column": "customer_id",
        "reason": "inconsistent_for_order_id",
        "identifier": "DH001",
    }


def test_csv_reader_accepts_utf8_sig() -> None:
    content = valid_frame().to_csv(index=False).encode("utf-8-sig")

    result = read_sales_file(file_name="sales.csv", content=content)

    assert list(result.columns) == REQUIRED_COLUMNS
    assert len(result) == 1


def test_xlsx_reader_accepts_first_sheet() -> None:
    buffer = BytesIO()
    valid_frame().to_excel(buffer, index=False, engine="openpyxl")

    result = read_sales_file(file_name="sales.xlsx", content=buffer.getvalue())

    assert list(result.columns) == REQUIRED_COLUMNS
    assert len(result) == 1


def test_reader_rejects_unknown_extension() -> None:
    with pytest.raises(AppError) as error:
        read_sales_file(file_name="sales.xls", content=b"not-empty")

    assert_app_error(error, "INVALID_FILE_TYPE")


def test_xlsx_reader_rejects_excessive_uncompressed_archive(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    buffer = BytesIO()
    with ZipFile(buffer, "w", compression=ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", "x" * 32)
    monkeypatch.setattr(file_reader, "MAX_XLSX_UNCOMPRESSED_BYTES", 16)

    with pytest.raises(AppError) as error:
        read_sales_file(file_name="sales.xlsx", content=buffer.getvalue())

    app_error = assert_app_error(error, "INVALID_FILE_TYPE")
    assert app_error.details["reason"] == "xlsx_archive_too_large"


def test_xlsx_reader_rejects_too_many_archive_members(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    buffer = BytesIO()
    with ZipFile(buffer, "w", compression=ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", "content")
        archive.writestr("xl/workbook.xml", "workbook")
    monkeypatch.setattr(file_reader, "MAX_XLSX_ARCHIVE_FILES", 1)

    with pytest.raises(AppError) as error:
        read_sales_file(file_name="sales.xlsx", content=buffer.getvalue())

    app_error = assert_app_error(error, "INVALID_FILE_TYPE")
    assert app_error.details["reason"] == "xlsx_archive_too_many_files"


def test_xlsx_reader_rejects_archive_path_traversal() -> None:
    buffer = BytesIO()
    with ZipFile(buffer, "w", compression=ZIP_DEFLATED) as archive:
        archive.writestr("../outside.xml", "unsafe")

    with pytest.raises(AppError) as error:
        read_sales_file(file_name="sales.xlsx", content=buffer.getvalue())

    app_error = assert_app_error(error, "INVALID_FILE_TYPE")
    assert app_error.details["reason"] == "unsafe_xlsx_archive"


def test_xlsx_reader_rejects_merged_cells() -> None:
    buffer = BytesIO()
    frame = valid_frame()
    frame.to_excel(buffer, index=False, engine="openpyxl")
    buffer.seek(0)

    workbook = load_workbook(buffer)
    workbook.active.merge_cells("A2:A3")
    merged = BytesIO()
    workbook.save(merged)
    workbook.close()

    with pytest.raises(AppError) as error:
        read_sales_file(file_name="sales.xlsx", content=merged.getvalue())

    app_error = assert_app_error(error, "INVALID_FILE_TYPE")
    assert app_error.details["reason"] == "merged_cells_not_supported"


def test_too_many_rows_is_rejected() -> None:
    frame = pd.concat([valid_frame(), valid_frame()], ignore_index=True)

    with pytest.raises(AppError) as error:
        validate_sales_data(frame, max_rows=1)

    assert_app_error(error, "TOO_MANY_ROWS")
